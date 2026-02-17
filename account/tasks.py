from io import BytesIO
from random import shuffle
import openpyxl

from PIL import Image, ImageDraw, ImageFont
from asgiref.sync import async_to_sync, sync_to_async
from celery.utils.log import get_task_logger
from channels.layers import get_channel_layer
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from django.core.mail import EmailMessage
from django.template.loader import render_to_string

from account.models import CustomUser
from facturation_backend.celery_conf import app
from facturation_backend.settings import STATIC_PATH
from facturation_backend.utils import ImageProcessor

logger = get_task_logger(__name__)


@app.task(bind=True, serializer="json", max_retries=3)
def send_email(self, user_pk, email_, mail_subject, message, code=None, type_=None):
    try:
        user = CustomUser.objects.get(pk=user_pk)
        email = EmailMessage(mail_subject, message, to=(email_,))
        email.content_subtype = "html"
        email.send(fail_silently=False)
        if type_ == "password_reset_code" and code is not None:
            user.password_reset_code = code
            user.save(update_fields=["password_reset_code"])
    except ObjectDoesNotExist:
        logger.error(f"Utilisateur {user_pk} introuvable pour la tâche d’e-mail")
        # Don't retry - user doesn't exist
        return
    except Exception as e:
        logger.error(f"Échec de l’envoi de l’e-mail pour l’utilisateur {user_pk} : {e}")
        raise self.retry(exc=e, countdown=60)  # Retry in 60 seconds


@app.task(bind=True, serializer="json", max_retries=3)
def send_csv_example_email(self, user_pk, email_):
    """Send import guide via email with CSV and Excel templates attached."""
    try:
        user = CustomUser.objects.get(pk=user_pk)
        
        # Create CSV content with headers only
        csv_headers = ("reference;type_article;designation;prix_achat;devise_prix_achat;"
                       "prix_vente;devise_prix_vente;tva;remarque;marque;categorie;emplacement;unite\n")
        
        # Create Excel template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"
        
        # Headers for Excel
        headers = [
            "reference", "type_article", "designation", "prix_achat", "devise_prix_achat",
            "prix_vente", "devise_prix_vente", "tva", "remarque", "marque", "categorie", 
            "emplacement", "unite"
        ]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save Excel to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        # Render HTML template
        mail_subject = "Guide d'importation des articles - Facturation"
        mail_template = "import_email_guide.html"
        message = render_to_string(
            mail_template,
            {
                "first_name": user.first_name,
            },
        )
        
        # Create email with HTML content
        email = EmailMessage(
            subject=mail_subject,
            body=message,
            to=(email_,)
        )
        email.content_subtype = "html"
        
        # Attach both CSV and Excel files
        email.attach("modele_articles.csv", csv_headers, "text/csv")
        email.attach("modele_articles.xlsx", excel_content, 
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Send email
        email.send(fail_silently=False)
        
    except ObjectDoesNotExist:
        logger.error(f"Utilisateur {user_pk} introuvable pour l'envoi du guide d'importation")
        return
    except Exception as e:
        logger.error(f"Échec de l'envoi du guide d'importation pour l'utilisateur {user_pk} : {e}")
        raise self.retry(exc=e, countdown=60)


@app.task(bind=True, serializer="json", max_retries=3)
def start_deleting_expired_codes(self, user_pk, type_):
    try:
        user = CustomUser.objects.get(pk=user_pk)
        if type_ == "password_reset":
            user.password_reset_code = None
            user.save(update_fields=["password_reset_code"])
    except ObjectDoesNotExist:
        logger.warning(
            f"Utilisateur {user_pk} introuvable pour start_deleting_expired_codes"
        )
        return
    except Exception as e:
        logger.error(
            f"Échec de la suppression des codes expirés pour l’utilisateur {user_pk} : {e}"
        )
        raise self.retry(exc=e, countdown=300)


# For generating Avatar
def random_color_picker():
    return [
        "#F3DCDC",
        "#FFD9A2",
        "#F8F2DA",
        "#DBF4EA",
        "#DBE8F4",
        "#D5CEEE",
        "#F3D8E1",
        "#EBD2AD",
        "#E2E4E2",
        "#FFA826",
        "#FED301",
        "#07CBAD",
        "#FF9DBF",
        "#CEB186",
        "#FF5D6B",
        "#0274D7",
        "#8669FB",
        "#878E88",
        "#0D070B",
    ]


def get_text_fill_color(bg_color):
    # white 255, 255, 255
    # black 0, 0, 0
    match bg_color:
        case (
            "#F3DCDC"
            | "#FFD9A2"
            | "#F8F2DA"
            | "#DBF4EA"
            | "#DBE8F4"
            | "#D5CEEE"
            | "#F3D8E1"
            | "#EBD2AD"
            | "#E2E4E2"
            | "#FFA826"
            | "#FED301"
            | "#07CBAD"
            | "#FF9DBF"
            | "#CEB186"
        ):
            return 0, 0, 0
        case "#FF5D6B" | "#0274D7" | "#8669FB" | "#878E88" | "#0D070B":
            return 255, 255, 255
        case _:
            # Return black color as default
            return 0, 0, 0


def from_img_to_io(image, format_):
    bytes_io = BytesIO()
    image.save(File(bytes_io), format=format_, save=True)
    bytes_io.seek(0)
    return bytes_io


def generate_avatar(last_name, first_name):
    colors = random_color_picker()
    shuffle(colors)
    color = colors.pop()
    fill = get_text_fill_color(color)
    avatar = Image.new("RGB", (600, 600), color=color)
    try:
        font_avatar = ImageFont.truetype(STATIC_PATH + "/fonts/Poppins-Bold.ttf", 240)
    except OSError:
        # Fallback to default PIL font if the custom font is missing
        font_avatar = ImageFont.load_default()
    drawn_avatar = ImageDraw.Draw(avatar)
    drawn_avatar.text(
        (100, 136), "{}.{}".format(first_name, last_name), font=font_avatar, fill=fill
    )
    return avatar


@app.task(bind=True, serializer="json", max_retries=3)
def generate_user_thumbnail(self, user_pk):
    try:
        user = CustomUser.objects.get(pk=user_pk)
        last_name = str(user.last_name[0]).upper() if user.last_name else "X"
        first_name = str(user.first_name[0]).upper() if user.first_name else "X"
        avatar = generate_avatar(last_name, first_name)
        avatar_ = from_img_to_io(avatar, "WEBP")
        user.save_image("avatar", avatar_)
        user.save_image("avatar_cropped", avatar_)
    except ObjectDoesNotExist:
        logger.error(f"Utilisateur {user_pk} introuvable pour la génération thumbnail")
        return
    except Exception as e:
        logger.error(f"Thumbnail génération échoué pour l'utilisateur {user_pk}: {e}")
        raise self.retry(exc=e, countdown=120)


def resize_images_v2(bytes_) -> BytesIO:
    image_processor = ImageProcessor()
    loaded_img = image_processor.load_image_from_io(bytes_)

    # Avatar 600x600 with blurred background
    avatar_img = image_processor.resize_with_blurred_background(loaded_img, 600)
    avatar_io = image_processor.from_img_to_io(avatar_img, "WEBP")

    return avatar_io


def generate_images_v2(query_, avatar: BytesIO):
    query_.save_image("avatar", avatar)


@app.task(bind=True, serializer="pickle", max_retries=3)
def resize_avatar(self, object_pk: int, avatar: BytesIO | None):
    try:
        user = CustomUser.objects.get(pk=object_pk)
        if not isinstance(avatar, BytesIO):
            return
        avatar_io = resize_images_v2(avatar)
        generate_images_v2(user, avatar_io)

        event = {
            "type": "receive_group_message",
            "message": {
                "type": "USER_AVATAR",
                "pk": user.pk,
                "avatar": user.get_absolute_avatar_img,
            },
        }
        channel_layer = get_channel_layer()
        async_send = sync_to_async(channel_layer.group_send)
        async_to_sync(async_send)(str(user.pk), event)
    except ObjectDoesNotExist:
        logger.error(
            f"Utilisateur {object_pk} introuvable pour le redimensionnement de l’avatar"
        )
        return
    except Exception as e:
        logger.error(
            f"Échec du redimensionnement de l’avatar pour l’utilisateur {object_pk} : {e}"
        )
        raise self.retry(exc=e, countdown=60)
